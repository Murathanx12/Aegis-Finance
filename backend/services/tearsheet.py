"""
Aegis Finance — One-page Tearsheets (HTML + Excel)
=====================================================

Bloomberg Green-Package-style one-page portfolio report. Two flavours:

  - render_portfolio_tearsheet_html(analysis) → full HTML document the
    frontend can download and print-to-PDF via the browser (no server
    PDF engine dependency).
  - render_portfolio_tearsheet_xlsx(analysis) → .xlsx bytes with sheets:
      - Summary (key metrics)
      - Holdings (ticker, weight, value, return, signal)
      - Risk (VaR/CVaR/Sharpe/Sortino + MCTR)
      - Factors (FF5 alpha, betas, style)
      - Stress (scenario impacts)

Input `analysis` is the dict returned by /api/portfolio/analyze — this
module is decoupled from the engine so it can be tested with fixtures.
"""

from __future__ import annotations

import io
import html
import logging
from datetime import datetime, timezone
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ── HTML tearsheet ───────────────────────────────────────────────────────────


def render_portfolio_tearsheet_html(
    analysis: dict,
    *,
    title: str = "Portfolio Tearsheet",
) -> str:
    """Return a self-contained HTML tearsheet, print-to-PDF friendly."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    rn = analysis.get("risk_number") or {}
    metrics = analysis.get("metrics") or {}
    holdings = analysis.get("holdings") or analysis.get("holdings_detail") or []
    factor = analysis.get("factor_exposures") or {}
    stress = analysis.get("stress_test") or {}
    attribution = analysis.get("attribution_summary") or {}
    mctr = (analysis.get("mctr_summary") or {}).get("top_risk_contributors") or []

    # Summary metrics table
    summary_rows = []
    summary_rows.append(("Risk Number (1-100)", _fmt(rn.get("risk_number"), "int")))
    summary_rows.append(("Risk Category", rn.get("category", "—")))
    summary_rows.append(("Annualised Return", _fmt(metrics.get("annual_return_pct"), "pct")))
    summary_rows.append(("Annualised Volatility", _fmt(metrics.get("annual_vol_pct"), "pct")))
    summary_rows.append(("Sharpe Ratio", _fmt(metrics.get("sharpe_ratio"), "num")))
    summary_rows.append(("Sortino Ratio", _fmt(metrics.get("sortino_ratio"), "num")))
    summary_rows.append(("Max Drawdown", _fmt(metrics.get("max_drawdown_pct"), "pct")))
    summary_rows.append(("VaR 95% (1-day)", _fmt(metrics.get("var_95_pct"), "pct")))
    summary_rows.append(("CVaR 95% (1-day)", _fmt(metrics.get("cvar_95_pct"), "pct")))
    if factor.get("alpha_annual") is not None:
        summary_rows.append(("FF5 Alpha (annual)", _fmt(factor["alpha_annual"] * 100, "pct")))
    if factor.get("market_beta") is not None:
        summary_rows.append(("Market Beta (FF5)", _fmt(factor["market_beta"], "num")))
    if factor.get("r_squared") is not None:
        summary_rows.append(("FF5 R²", _fmt(factor["r_squared"] * 100, "pct")))

    # Holdings table
    holdings_rows = []
    for h in holdings:
        holdings_rows.append(
            (
                h.get("ticker") or "—",
                _fmt(h.get("weight_pct") or (h.get("weight", 0) * 100), "pct"),
                _fmt(h.get("current_price"), "num"),
                _fmt(h.get("value"), "num"),
                _fmt(h.get("annual_return_pct"), "pct"),
                h.get("signal") or h.get("action") or "—",
            )
        )

    # Stress rows
    stress_rows = []
    for name, s in (stress.get("scenarios") or {}).items():
        stress_rows.append(
            (
                name,
                _fmt(s.get("portfolio_drawdown_pct"), "pct"),
                _fmt(s.get("sp500_drawdown_pct"), "pct"),
                s.get("relative_to_market") or "—",
            )
        )

    # MCTR rows
    mctr_rows = []
    for c in mctr[:10]:
        mctr_rows.append(
            (
                c.get("ticker") or "—",
                _fmt(c.get("weight_pct"), "pct"),
                _fmt(c.get("risk_contrib_pct"), "pct"),
                _fmt(c.get("mctr"), "num"),
            )
        )

    return _TEMPLATE.format(
        title=html.escape(title),
        generated_at=now,
        summary_rows=_html_table(
            ["Metric", "Value"], summary_rows,
        ),
        holdings_rows=_html_table(
            ["Ticker", "Weight", "Price", "Value", "1Y Return", "Signal"],
            holdings_rows,
        ) if holdings_rows else "<p class='empty'>No holdings data.</p>",
        stress_rows=_html_table(
            ["Scenario", "Portfolio DD", "SP500 DD", "Relative"],
            stress_rows,
        ) if stress_rows else "<p class='empty'>No stress test data.</p>",
        mctr_rows=_html_table(
            ["Ticker", "Weight", "Risk Contrib.", "MCTR"],
            mctr_rows,
        ) if mctr_rows else "<p class='empty'>No risk decomposition available.</p>",
        attribution_block=_render_attribution(attribution),
        brinson_note=(
            f"Period: {html.escape(str(attribution.get('period') or 'n/a'))}"
            if attribution else ""
        ),
    )


def _render_attribution(attr: dict) -> str:
    if not attr:
        return ""
    fields = [
        ("Portfolio Return", attr.get("portfolio_return"), "pct"),
        ("Benchmark Return", attr.get("benchmark_return"), "pct"),
        ("Active Return", attr.get("total_active_return"), "pct"),
        ("Allocation Effect", attr.get("total_allocation_effect"), "pct"),
        ("Selection Effect", attr.get("total_selection_effect"), "pct"),
        ("Interaction Effect", attr.get("total_interaction_effect"), "pct"),
    ]
    rows = [(name, _fmt(val, unit)) for name, val, unit in fields if val is not None]
    return _html_table(["Brinson Effect", "Value"], rows) if rows else ""


def _html_table(headers: list[str], rows: list[tuple]) -> str:
    thead = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
    tbody = ""
    for row in rows:
        tbody += "<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in row) + "</tr>"
    return f"<table><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>"


def _fmt(v, kind: str = "num") -> str:
    if v is None:
        return "—"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if kind == "pct":
        return f"{v:+.2f}%"
    if kind == "int":
        return f"{int(round(v))}"
    # "num"
    if abs(v) >= 1_000_000:
        return f"{v:,.0f}"
    return f"{v:,.2f}"


_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{title} — Aegis Finance</title>
  <style>
    :root {{
      --bg: #fff; --fg: #111; --muted: #555; --border: #e5e5e5;
      --accent: #b88a00;
    }}
    @media (prefers-color-scheme: dark) {{
      :root {{ --bg: #0a0a0a; --fg: #e5e5e5; --muted: #9a9a9a; --border: #222; }}
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0; padding: 0; background: var(--bg); color: var(--fg);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }}
    .page {{ max-width: 960px; margin: 0 auto; padding: 32px; }}
    header {{
      display: flex; align-items: baseline; justify-content: space-between;
      border-bottom: 1px solid var(--border); padding-bottom: 16px; margin-bottom: 24px;
    }}
    h1 {{ font-size: 22px; margin: 0; font-weight: 700; }}
    h2 {{ font-size: 14px; margin: 28px 0 10px; text-transform: uppercase; letter-spacing: 0.12em; color: var(--muted); }}
    .meta {{ color: var(--muted); font-size: 12px; font-family: monospace; }}
    .badge {{ font-family: monospace; font-size: 10px; letter-spacing: 0.1em; color: var(--accent); text-transform: uppercase; }}
    table {{
      width: 100%; border-collapse: collapse; font-size: 13px;
    }}
    th, td {{
      padding: 8px 10px; text-align: left; border-bottom: 1px solid var(--border);
    }}
    th {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; color: var(--muted); font-weight: 600; }}
    td {{ font-variant-numeric: tabular-nums; }}
    td:nth-child(n+2) {{ text-align: right; font-family: monospace; }}
    .empty {{ color: var(--muted); font-size: 12px; font-style: italic; }}
    .cols {{ display: grid; grid-template-columns: 1fr 1fr; gap: 32px; }}
    .footer {{
      margin-top: 40px; padding-top: 16px; border-top: 1px solid var(--border);
      font-size: 11px; color: var(--muted);
    }}
    @media print {{
      body {{ background: #fff; color: #000; }}
      .page {{ padding: 0; }}
      .no-print {{ display: none; }}
    }}
  </style>
</head>
<body>
<div class="page">
  <header>
    <div>
      <div class="badge">Aegis Finance · Portfolio Tearsheet</div>
      <h1>{title}</h1>
    </div>
    <div class="meta">Generated {generated_at}</div>
  </header>

  <h2>Summary</h2>
  {summary_rows}

  <div class="cols">
    <div>
      <h2>Holdings</h2>
      {holdings_rows}
    </div>
    <div>
      <h2>Top Risk Contributors (MCTR)</h2>
      {mctr_rows}
    </div>
  </div>

  <h2>Historical Stress Tests</h2>
  {stress_rows}

  {attribution_block}
  <p class="meta">{brinson_note}</p>

  <div class="footer">
    Educational tool only. Not financial advice. All figures are model-based estimates.
  </div>
</div>
</body>
</html>
"""


# ── Excel export ─────────────────────────────────────────────────────────────


def render_portfolio_tearsheet_xlsx(analysis: dict) -> bytes:
    """Return an .xlsx workbook as bytes. Multi-sheet Bloomberg PORT-style."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("openpyxl not installed — cannot generate .xlsx") from e

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    _write_summary_sheet(ws, analysis, Font, PatternFill, Alignment)

    # Holdings
    ws = wb.create_sheet("Holdings")
    _write_holdings_sheet(ws, analysis, Font, PatternFill, Alignment)

    # Risk / MCTR
    ws = wb.create_sheet("Risk")
    _write_risk_sheet(ws, analysis, Font, PatternFill, Alignment)

    # Factors
    ws = wb.create_sheet("Factors")
    _write_factors_sheet(ws, analysis, Font, PatternFill, Alignment)

    # Stress
    ws = wb.create_sheet("Stress Tests")
    _write_stress_sheet(ws, analysis, Font, PatternFill, Alignment)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _header(ws, row, cols, Font, PatternFill, Alignment):
    fill = PatternFill("solid", fgColor="2D2D2D")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="left", vertical="center")
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _write_summary_sheet(ws, analysis, Font, PatternFill, Alignment):
    ws["A1"] = "Aegis Finance Portfolio Tearsheet"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {datetime.now(timezone.utc).isoformat(timespec='seconds')}"
    ws["A2"].font = Font(italic=True, color="808080")

    rn = analysis.get("risk_number") or {}
    metrics = analysis.get("metrics") or {}
    factor = analysis.get("factor_exposures") or {}

    rows = [
        ("Risk Number (1-100)", rn.get("risk_number")),
        ("Risk Category", rn.get("category")),
        ("Annualised Return (%)", metrics.get("annual_return_pct")),
        ("Annualised Volatility (%)", metrics.get("annual_vol_pct")),
        ("Sharpe Ratio", metrics.get("sharpe_ratio")),
        ("Sortino Ratio", metrics.get("sortino_ratio")),
        ("Max Drawdown (%)", metrics.get("max_drawdown_pct")),
        ("VaR 95% 1d (%)", metrics.get("var_95_pct")),
        ("CVaR 95% 1d (%)", metrics.get("cvar_95_pct")),
        ("FF5 Alpha annual (%)", _times_hundred(factor.get("alpha_annual"))),
        ("Market Beta", factor.get("market_beta")),
        ("FF5 R² (%)", _times_hundred(factor.get("r_squared"))),
    ]
    _header(ws, 4, ["Metric", "Value"], Font, PatternFill, Alignment)
    for i, (k, v) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v if v is not None else "—")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _write_holdings_sheet(ws, analysis, Font, PatternFill, Alignment):
    holdings = analysis.get("holdings") or analysis.get("holdings_detail") or []
    _header(
        ws,
        1,
        ["Ticker", "Shares", "Price", "Value", "Weight %", "1Y Return %", "Signal"],
        Font,
        PatternFill,
        Alignment,
    )
    for i, h in enumerate(holdings, start=2):
        weight = h.get("weight_pct")
        if weight is None and h.get("weight") is not None:
            weight = h["weight"] * 100
        ws.cell(row=i, column=1, value=h.get("ticker"))
        ws.cell(row=i, column=2, value=h.get("shares"))
        ws.cell(row=i, column=3, value=h.get("current_price"))
        ws.cell(row=i, column=4, value=h.get("value"))
        ws.cell(row=i, column=5, value=weight)
        ws.cell(row=i, column=6, value=h.get("annual_return_pct"))
        ws.cell(row=i, column=7, value=h.get("signal") or h.get("action"))
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 14


def _write_risk_sheet(ws, analysis, Font, PatternFill, Alignment):
    ws["A1"] = "Risk Decomposition (MCTR)"
    ws["A1"].font = Font(bold=True, size=12)
    mctr = (analysis.get("mctr_summary") or {}).get("top_risk_contributors") or []
    _header(
        ws,
        3,
        ["Ticker", "Weight %", "Risk Contribution %", "MCTR"],
        Font,
        PatternFill,
        Alignment,
    )
    for i, c in enumerate(mctr, start=4):
        ws.cell(row=i, column=1, value=c.get("ticker"))
        ws.cell(row=i, column=2, value=c.get("weight_pct"))
        ws.cell(row=i, column=3, value=c.get("risk_contrib_pct"))
        ws.cell(row=i, column=4, value=c.get("mctr"))
    for col in "ABCD":
        ws.column_dimensions[col].width = 18


def _write_factors_sheet(ws, analysis, Font, PatternFill, Alignment):
    factor = analysis.get("factor_exposures") or {}
    ws["A1"] = "Fama-French 5-Factor Exposures"
    ws["A1"].font = Font(bold=True, size=12)
    _header(ws, 3, ["Metric", "Value"], Font, PatternFill, Alignment)
    rows = [
        ("Alpha (annual)", _times_hundred(factor.get("alpha_annual"))),
        ("Market Beta", factor.get("market_beta")),
        ("R²", _times_hundred(factor.get("r_squared"))),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v if v is not None else "—")
    # Per-stock betas
    stocks = factor.get("stocks") or {}
    if stocks:
        start = 4 + len(rows) + 2
        ws.cell(row=start - 1, column=1, value="Per-holding betas").font = Font(bold=True)
        _header(ws, start, ["Ticker", "Market Beta", "Style"], Font, PatternFill, Alignment)
        for i, (t, s) in enumerate(stocks.items(), start=start + 1):
            ws.cell(row=i, column=1, value=t)
            ws.cell(row=i, column=2, value=s.get("market_beta"))
            ws.cell(row=i, column=3, value=s.get("style"))
    for col in "ABC":
        ws.column_dimensions[col].width = 24


def _write_stress_sheet(ws, analysis, Font, PatternFill, Alignment):
    stress = analysis.get("stress_test") or {}
    ws["A1"] = "Historical Stress Tests"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Worst case: {stress.get('worst_scenario') or '—'}"
    _header(
        ws,
        4,
        ["Scenario", "Portfolio DD %", "SP500 DD %", "Relative"],
        Font,
        PatternFill,
        Alignment,
    )
    for i, (name, s) in enumerate((stress.get("scenarios") or {}).items(), start=5):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=s.get("portfolio_drawdown_pct"))
        ws.cell(row=i, column=3, value=s.get("sp500_drawdown_pct"))
        ws.cell(row=i, column=4, value=s.get("relative_to_market"))
    for col in "ABCD":
        ws.column_dimensions[col].width = 24


def _times_hundred(v: Optional[Any]) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v) * 100
    except (TypeError, ValueError):
        return None
