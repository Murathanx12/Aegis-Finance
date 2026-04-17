"""Tests for HTML + Excel tearsheet rendering."""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient


SAMPLE_ANALYSIS = {
    "risk_number": {"risk_number": 62, "category": "moderate-aggressive"},
    "metrics": {
        "annual_return_pct": 11.4,
        "annual_vol_pct": 14.2,
        "sharpe_ratio": 0.80,
        "sortino_ratio": 1.15,
        "max_drawdown_pct": -18.3,
        "var_95_pct": -1.9,
        "cvar_95_pct": -2.6,
    },
    "holdings": [
        {
            "ticker": "AAPL",
            "shares": 10,
            "current_price": 230.0,
            "value": 2300.0,
            "weight_pct": 40.0,
            "annual_return_pct": 22.1,
            "signal": "Buy",
        },
        {
            "ticker": "SPY",
            "shares": 8,
            "current_price": 550.0,
            "value": 4400.0,
            "weight_pct": 60.0,
            "annual_return_pct": 12.3,
            "signal": "Hold",
        },
    ],
    "factor_exposures": {
        "alpha_annual": 0.015,
        "market_beta": 0.92,
        "r_squared": 0.88,
        "stocks": {
            "AAPL": {"market_beta": 1.15, "style": "Large Growth"},
            "SPY": {"market_beta": 1.00, "style": "Large Blend"},
        },
    },
    "stress_test": {
        "worst_scenario": "GFC 2008",
        "scenarios": {
            "GFC 2008": {
                "portfolio_drawdown_pct": -38.1,
                "sp500_drawdown_pct": -55.0,
                "relative_to_market": "outperformed",
            },
            "COVID 2020": {
                "portfolio_drawdown_pct": -22.4,
                "sp500_drawdown_pct": -34.0,
                "relative_to_market": "outperformed",
            },
        },
    },
    "attribution_summary": {
        "period": "1mo",
        "portfolio_return": 2.1,
        "benchmark_return": 1.6,
        "total_active_return": 0.5,
        "total_allocation_effect": 0.1,
        "total_selection_effect": 0.3,
        "total_interaction_effect": 0.1,
    },
    "mctr_summary": {
        "top_risk_contributors": [
            {"ticker": "AAPL", "weight_pct": 40.0, "risk_contrib_pct": 48.0, "mctr": 0.0225},
            {"ticker": "SPY", "weight_pct": 60.0, "risk_contrib_pct": 52.0, "mctr": 0.0162},
        ],
    },
}


# ── HTML tearsheet ───────────────────────────────────────────────────────────


class TestHtmlRenderer:
    def test_renders_complete_document(self):
        from backend.services.tearsheet import render_portfolio_tearsheet_html

        out = render_portfolio_tearsheet_html(SAMPLE_ANALYSIS, title="My Port")
        assert out.startswith("<!doctype html>")
        # Key fields make it in
        assert "My Port" in out
        assert "AAPL" in out
        assert "SPY" in out
        assert "GFC 2008" in out
        assert "FF5 Alpha" in out
        # Tabular numeric formatting
        assert "62" in out  # risk number
        # Attribution section
        assert "Active Return" in out

    def test_handles_empty_analysis(self):
        from backend.services.tearsheet import render_portfolio_tearsheet_html

        out = render_portfolio_tearsheet_html({})
        assert "<!doctype html>" in out
        # Must not blow up — missing sections surfaced as "empty"
        assert "No holdings" in out or "—" in out

    def test_escapes_html_in_ticker(self):
        analysis = {**SAMPLE_ANALYSIS, "holdings": [{"ticker": "<script>", "weight_pct": 100}]}
        from backend.services.tearsheet import render_portfolio_tearsheet_html
        out = render_portfolio_tearsheet_html(analysis)
        assert "<script>" not in out.split("</head>")[1]  # no raw tag in body
        assert "&lt;script&gt;" in out


# ── Excel ────────────────────────────────────────────────────────────────────


class TestXlsxRenderer:
    def test_produces_valid_xlsx_bytes(self):
        from backend.services.tearsheet import render_portfolio_tearsheet_xlsx
        import openpyxl

        blob = render_portfolio_tearsheet_xlsx(SAMPLE_ANALYSIS)
        assert isinstance(blob, bytes)
        assert len(blob) > 1000
        # Round-trip through openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        assert "Summary" in wb.sheetnames
        assert "Holdings" in wb.sheetnames
        assert "Risk" in wb.sheetnames
        assert "Factors" in wb.sheetnames
        assert "Stress Tests" in wb.sheetnames

    def test_holdings_sheet_has_rows(self):
        from backend.services.tearsheet import render_portfolio_tearsheet_xlsx
        import openpyxl

        blob = render_portfolio_tearsheet_xlsx(SAMPLE_ANALYSIS)
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        ws = wb["Holdings"]
        # Header in row 1, AAPL in row 2
        assert ws.cell(row=2, column=1).value == "AAPL"
        assert ws.cell(row=3, column=1).value == "SPY"

    def test_summary_sheet_keys(self):
        from backend.services.tearsheet import render_portfolio_tearsheet_xlsx
        import openpyxl

        blob = render_portfolio_tearsheet_xlsx(SAMPLE_ANALYSIS)
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        ws = wb["Summary"]
        metric_col = [ws.cell(row=i, column=1).value for i in range(5, 20)]
        assert "Sharpe Ratio" in metric_col
        assert "Max Drawdown (%)" in metric_col


# ── Endpoint integration ─────────────────────────────────────────────────────


@pytest.fixture
def client():
    from backend.main import app
    return TestClient(app)


class TestTearsheetEndpoints:
    def test_html_endpoint(self, client, monkeypatch):
        from backend.routers import portfolio

        monkeypatch.setattr(portfolio, "_analyze_with_risk_number", lambda h: SAMPLE_ANALYSIS)

        r = client.post(
            "/api/portfolio/tearsheet.html",
            json={
                "holdings": [{"ticker": "AAPL", "shares": 10, "current_price": 230}],
                "title": "My Ports",
            },
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/html")
        assert "My Ports" in r.text
        assert "AAPL" in r.text

    def test_xlsx_endpoint(self, client, monkeypatch):
        from backend.routers import portfolio

        monkeypatch.setattr(portfolio, "_analyze_with_risk_number", lambda h: SAMPLE_ANALYSIS)
        r = client.post(
            "/api/portfolio/tearsheet.xlsx",
            json={
                "holdings": [{"ticker": "AAPL", "shares": 10, "current_price": 230}],
                "title": "XL Port",
            },
        )
        assert r.status_code == 200
        assert "spreadsheet" in r.headers["content-type"]
        assert "attachment" in r.headers.get("content-disposition", "")
        assert r.content[:2] == b"PK"  # zip/xlsx magic bytes
